#!/usr/bin/env python3
# extract auditor invoices in various formS of files and write to a CSV

import os
from invoice2data import extract_data
from invoice2data.extract.loader import read_templates
import csv
import datetime
import xlrd
import openpyxl
import docx
import re
import pandas as pd

os.system("cls")

PASSWORD = "Silliker7"


def getCleanString(anystring: str) -> str:
    """rid strings of /t,space and #"""
    cleanedstring = str(anystring).strip().replace("#", "")
    return cleanedstring


def getText(filename):
    """get full text from a DOCX file"""
    doc = docx.Document(filename)
    fullText = []
    for para in doc.paragraphs:
        fullText.append(para.text)
    return "\n".join(fullText)


def turnDTintoSTR(somedatetimeobject):
    """turn datetime object into string"""
    if isinstance(somedatetimeobject, list):
        newlistwithoutnone = [i for i in somedatetimeobject if i is not None]
        if len(newlistwithoutnone) >= 1:
            somedatetimeobject = min(newlistwithoutnone)
            somedatetimeobject = str(
                str(somedatetimeobject.day).rjust(2, "0")
                + str(somedatetimeobject.month).rjust(2, "0")
                + str(somedatetimeobject.year)[2:4]
            )
        else:
            somedatetimeobject = ""
    elif isinstance(somedatetimeobject, datetime.date):
        somedatetimeobject = str(
            str(somedatetimeobject.day).rjust(2, "0")
            + str(somedatetimeobject.month).rjust(2, "0")
            + str(somedatetimeobject.year)[2:4]
        )
    return somedatetimeobject


def writeToCSV(
    vendorname=None,
    vendornumber=None,
    invoicenumber=None,
    invoicetotal=None,
    invoicedate=None,
    desc=None,
    today=None,
):
    """write invoice infor into csv"""
    outputdict = {
        "Vendor Name": vendorname,
        "Vendor Number": vendornumber,
        "Inv Number": invoicenumber,
        "Inv Total": invoicetotal,
        "Inv Date": invoicedate,
        "Currency": "AUD",
        "Desc": desc,
        "Process Date": today,
    }
    with open("./auditorinvoicedata.csv", "a", newline="") as f:
        w = csv.writer(f)
        w.writerow(outputdict.values())


def getTodayInString():
    """get today's date in a uniform string format"""
    today = datetime.datetime.now()
    today = str(today.day).zfill(2) + str(today.month).zfill(2) + str(today.year)
    return today


def sanitize_filename(filename: str) -> str:
    """
    Sanitize a filename by replacing characters not allowed in filenames with underscores.
    """
    forbidden_chars = '<>:"/\\|?*'
    sanitized_filename = "".join(
        "_" if char in forbidden_chars else char for char in filename
    )
    return sanitized_filename


def scrapePDFAuditorInvoice(filepath):
    """use invoice2data module to extra PDF invoice info"""
    # templates = read_templates('Invoice2data/Template/')
    templates = read_templates(
        ".\Dropbox\Side Hussle\Python\Finished Projects\Invoice2data\Template"
    )
    today = datetime.datetime.now()
    today = str(today.day).zfill(2) + str(today.month).zfill(2) + str(today.year)
    result = extract_data(filepath, templates=templates)
    result["register_date"] = today  # add time stamp
    result["date"] = str(turnDTintoSTR(result["date"]))

    with open("auditorinvoicedata.csv", "a", newline="") as f:
        w = csv.writer(f)
        w.writerow(result.values())

    os.rename(
        filepath,
        "Invoice/"
        + sanitize_filename(
            str(result["issuer"] + " " + result["invoice_number"] + ".pdf")
        ),
    )


def scrapeDOCXAuditorInvoice(docxfilepath):
    """currently only 1 vendor use DOCX format---KAR FOOD"""
    whole_docx_text = getText(docxfilepath)

    # build re object
    vendor_name_re = re.compile(r"(KAR FOOD SAFETY AUDITING PTY LTD)")
    invoice_number_re = re.compile(r"Invoice No:\s+(\d+)")
    invoice_total_re = re.compile(r"\nTotal\s+\$\s+(\d+\.\d{2})")
    invoice_date_re = re.compile(r"Date:\s+(\d{1,2}/\d{2}/\d{2,4})")

    # search re object
    if vendor_name_re.search(whole_docx_text) is not None:
        vendorname = vendor_name_re.search(whole_docx_text).group(1)
        vendornumber = 504900
    else:
        vendorname = ""
        vendornumber = ""

    if invoice_number_re.search(whole_docx_text) is not None:
        invoicenumber = invoice_number_re.search(whole_docx_text).group(1)
    else:
        invoicenumber = ""

    if invoice_total_re.search(whole_docx_text) is not None:
        invoicetotal = invoice_total_re.search(whole_docx_text).group(1)
    else:
        invoicetotal = ""

    if invoice_date_re.search(whole_docx_text) is not None:
        invoicedate = invoice_date_re.search(whole_docx_text).group(1)
    else:
        invoicedate = ""

    today = getTodayInString()
    writeToCSV(
        vendorname,
        vendornumber,
        invoicenumber,
        invoicetotal,
        invoicedate,
        "DOCX",
        today,
    )

    os.rename(
        docxfilepath,
        "Invoice/"
        + sanitize_filename(str(vendorname) + " " + str(invoicenumber) + ".docx"),
    )


def scrapeXLSAuditorInvoice(filename):
    """XLS format is major source of invoices"""
    wb = xlrd.open_workbook("./Invoice/" + filename)
    for sh in wb.sheets():
        if sh.name == "Auditor Invoice" or sh.name == "File Review Invoice":
            for row in range(sh.nrows):
                for col in range(sh.ncols):
                    myCell = sh.cell(row, col)
                    if myCell.value == "Invoice Total":
                        invoicetotal = sh.cell(row, col + 3).value
                        invoicetotal = round(invoicetotal, 2)
                    elif myCell.value == "Inv No":
                        invoicenumber = sh.cell(row, col + 1).value
                        invoicenumber = getCleanString(invoicenumber)
                    elif myCell.value == "Inv Date":
                        invoicedate = sh.cell(row, col + 1).value
                        if type(invoicedate) is float:
                            invoicedate = xlrd.xldate_as_datetime(invoicedate, 0)
                            invoicedate = invoicedate.date()
                    elif myCell.value == "Vendor Number":
                        vendornumber = sh.cell(row, col + 1).value
                        vendornumber = getCleanString(vendornumber)
                    elif myCell.value == "Auditor Company Name":
                        vendorname = sh.cell(row, col + 1).value

            today = getTodayInString()
            writeToCSV(
                vendorname,
                vendornumber,
                invoicenumber,
                invoicetotal,
                invoicedate,
                "XLS",
                today,
            )
            try:
                os.rename(
                    "Invoice/" + filename,
                    "Invoice/"
                    + sanitize_filename(
                        str(vendorname) + " " + str(invoicenumber) + ".xls"
                    )
                )
            except FileNotFoundError:
                print(f"The file '{filename}' does not exist.")
            except PermissionError:
                print(f"You do not have permission to rename '{filename}'.")
            except OSError as e:
                print(f"An error occurred while renaming '{filename}' : {e}")
            except Exception as e:
                print(f"An unexpected error occurred: {e}")


def scrapeXLSXAuditorInvoice(wb):
    """only 1 vendor use XLSX format invoices"""
    wrb = openpyxl.open("./Invoice/" + wb, read_only=True, data_only=True)
    sh = wrb["Auditor invoice"]
    for row in range(1, sh.max_row + 1):
        for col in range(1, sh.max_column + 1):
            myCell = sh.cell(row, col)
            if myCell.value == "Invoice Total":
                invoicetotal = sh.cell(row, col + 2).value
            elif myCell.value == "Inv No":
                invoicenumber = sh.cell(row, col + 1).value
                invoicenumber = getCleanString(invoicenumber)
            elif myCell.value == "Inv Date":
                invoicedate = sh.cell(row, col + 1).value
                invoicedate = invoicedate.strftime("%d%m%y")
            elif myCell.value == "3532-13706":
                vendornumber = 506930
                vendorname = "Foodsafe Focus"

    today = getTodayInString()
    writeToCSV(
        vendorname,
        vendornumber,
        invoicenumber,
        invoicetotal,
        invoicedate,
        "XLSX",
        today,
    )

    os.rename(
        "./Invoice/" + wb,
        "./Invoice/"
        + sanitize_filename(str(vendorname) + " " + str(invoicenumber) + ".xlsx"),
    )


def check_duplicated_invoice_number():
    # Read in the Excel table
    df = pd.read_csv("auditorinvoicedata.csv", engine="python")

    # Slice the table based on date in column H
    today = getTodayInString()
    df_slice = df[df["process_date"] == int(today)].drop_duplicates()

    # Loop through each row in the slice
    for index, row in df_slice.iterrows():
        # Check if Column A AND Column C has any same value in the slice or the whole dataframe
        if (
            df[
                (
                    (df["invoice_number"] == row["invoice_number"])
                )
                & (df["process_date"] != int(today))
            ].shape[0]
            > 0
        ):
            # If there is a match, warn the user
            print(
                f"\nWarning: Duplicate Invoice Number. Row {index+2};From vendor: "
                + df_slice.loc[index, "company"]
                + "; Invoice Number: "
                + df_slice.loc[index, "invoice_number"]
                + "\n\n"
            )


def walkfiles():
    unprocessed_files = []
    for subdir, dirs, files in os.walk("./Invoice"):
        for filename in files:
            if filename.endswith(".xls"):
                try:
                    scrapeXLSAuditorInvoice(filename)
                except:
                    unprocessed_files.append(filename)
                    continue
            elif filename.endswith(".xlsx"):
                try:
                    scrapeXLSXAuditorInvoice(filename)
                except:
                    unprocessed_files.append(filename)
                    continue
            elif filename.endswith(".pdf"):
                try:
                    path = "Invoice/" + filename
                    scrapePDFAuditorInvoice(path)
                except Exception as e:
                    unprocessed_files.append(filename)
                    print(f"Error: {e}")
                    continue
            elif filename.endswith(".docx"):
                try:
                    docxpath = "Invoice/" + filename
                    scrapeDOCXAuditorInvoice(docxpath)
                except:
                    unprocessed_files.append(filename)
                    continue
            else:
                unprocessed_files.append(filename)
                continue
    if len(unprocessed_files) > 0:
        print("\n\nThe Following Invoices are Not Processed:\n")
        print("\n".join(map(str, unprocessed_files)))
    else:
        print("\n\nAll Invoices Info Extracted!\n")


# close invoicedata.csv if it's already opened
try:
    os.rename("auditorinvoicedata.csv", "auditorinvoicedata.csv")
except IOError:
    input("Please Close auditornvoicedata.csv")
    quit()
else:
    walkfiles()
    check_duplicated_invoice_number()
